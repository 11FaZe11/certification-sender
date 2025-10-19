import io
from openpyxl import load_workbook
import os
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from typing import Tuple, Optional
from matplotlib import font_manager
import shutil
from urllib.parse import quote


def ensure_fonts_dir(fonts_dir: str = "fonts"):
    """Create the fonts directory if it doesn't exist."""
    if not os.path.exists(fonts_dir):
        os.makedirs(fonts_dir)


def register_font_file(font_path: str) -> Optional[str]:
    """Register a TTF font file with reportlab and return the registered font name.
    Ensures uniqueness if the name is already registered.
    """
    try:
        font_name = os.path.splitext(os.path.basename(font_path))[0]
        base_font_name = font_name
        suffix = 1
        while font_name in pdfmetrics.getRegisteredFontNames():
            font_name = f"{base_font_name}_{suffix}"
            suffix += 1
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        return font_name
    except Exception:
        return None


def download_google_font(font_family: str, fonts_dir: str = "fonts") -> str:
    """Attempt to download a TTF for the requested font family from the google/fonts GitHub repository.
    On success returns the registered ReportLab font name; otherwise returns an empty string.
    """
    # Local import so the module is optional at static-check time
    try:
        import requests
    except Exception:
        print("The 'requests' package is not installed. Install it (pip install -r requirements.txt) to enable automatic font downloads.")
        return ""

    ensure_fonts_dir(fonts_dir)
    # Normalize family names for URL building
    family_no_space = font_family.replace(" ", "").lower()
    family_dash = font_family.replace(" ", "-").lower()
    candidates = []

    # Common filename patterns to try
    base_names = [
        f"{font_family}-Regular.ttf",
        f"{font_family.title().replace(' ', '')}-Regular.ttf",
        f"{font_family.replace(' ', '')}-Regular.ttf",
        f"{family_no_space}-regular.ttf",
        f"{family_dash}-regular.ttf",
        f"{font_family}.ttf",
        f"{font_family.title()}.ttf",
        f"{font_family.replace(' ', '')}.ttf",
    ]

    # Try both ofl and apache folders
    owners = ["ofl", "apache"]
    for owner in owners:
        for name in base_names:
            candidates.append(f"https://github.com/google/fonts/raw/main/{owner}/{family_dash}/{quote(name)}")
            candidates.append(f"https://github.com/google/fonts/raw/main/{owner}/{family_no_space}/{quote(name)}")
            candidates.append(f"https://github.com/google/fonts/raw/main/{owner}/{font_family.replace(' ', '').lower()}/{quote(name)}")

    # Also try direct family folder names with title casing (many Google fonts folders are the family name in lower-case)
    candidates.append(f"https://github.com/google/fonts/raw/main/ofl/{family_dash}/{quote(font_family.title().replace(' ', '') + '-Regular.ttf')}")

    for url in candidates:
        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200 and resp.content:
                # Heuristic: size should be > 1KB
                if len(resp.content) < 1024:
                    continue
                # Save file using the last part of the URL
                filename = os.path.basename(url.split('?')[0])
                # Some filenames are percent-encoded; unquote filename
                try:
                    from urllib.parse import unquote
                    filename = unquote(filename)
                except Exception:
                    pass
                local_path = os.path.join(fonts_dir, filename)
                with open(local_path, 'wb') as f:
                    f.write(resp.content)
                # Try to register immediately; if registration fails, continue searching
                registered = register_font_file(local_path)
                if registered:
                    print(f"Downloaded and registered font '{font_family}' as '{registered}' from {url}")
                    return registered
                else:
                    # remove file if registration failed
                    try:
                        os.remove(local_path)
                    except Exception:
                        pass
        except Exception:
            # ignore network errors for individual URLs and keep trying
            continue

    # If we couldn't find via GitHub, try the Google Fonts CSS endpoint to locate woff2 resources (best-effort)
    try:
        css_url = f"https://fonts.googleapis.com/css2?family={quote(font_family)}"
        resp = requests.get(css_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code == 200 and "url(" in resp.text:
            # Find first url(...) occurrence
            import re
            urls = re.findall(r"url\((https:[^)]+)\)", resp.text)
            for found in urls:
                # Try to fetch and save — many are woff2; reportlab needs TTF, so skip woff2
                if found.endswith('.ttf'):
                    try:
                        rr = requests.get(found, timeout=10)
                        if rr.status_code == 200 and len(rr.content) > 1024:
                            filename = os.path.basename(found.split('?')[0])
                            local_path = os.path.join(fonts_dir, filename)
                            with open(local_path, 'wb') as f:
                                f.write(rr.content)
                            registered = register_font_file(local_path)
                            if registered:
                                print(f"Downloaded and registered font '{font_family}' as '{registered}' from {found}")
                                return registered
                    except Exception:
                        continue
    except Exception:
        pass

    return ""


def get_system_fonts(fonts_dir: str = "fonts"):
    """
    Finds and registers all system TrueType fonts with reportlab
    and returns a list of their names. Also registers fonts placed in the local `fonts/` folder.
    """
    font_paths = font_manager.findSystemFonts(fontpaths=None, fontext='ttf')
    font_names = []
    # Register system fonts
    for font_path in font_paths:
        try:
            registered = register_font_file(font_path)
            if registered:
                font_names.append(registered)
        except Exception:
            pass

    # Also register fonts from the project's fonts directory (if any)
    if os.path.isdir(fonts_dir):
        for root, _, files in os.walk(fonts_dir):
            for f in files:
                if f.lower().endswith('.ttf'):
                    path = os.path.join(root, f)
                    try:
                        registered = register_font_file(path)
                        if registered:
                            font_names.append(registered)
                    except Exception:
                        pass

    return sorted(list(set(font_names)))

def col_letter_to_index(letter):
    try:
        letter = letter.strip().upper()
        if not letter.isalpha():
            raise ValueError("Column letter must be alphabetic.")
        index = 0
        for char in letter:
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index
    except Exception as e:
        print(f"Invalid column letter '{letter}': {e}")
        return None

def create_overlay_stream(text: str, width: float, height: float, x: float, y: float,
                          font_name: str = "Helvetica", font_size: int = 24,
                          color_rgb: Tuple[float, float, float] = (0, 0, 0)) -> io.BytesIO:
    """Create a PDF in-memory with the given text at (x, y) using reportlab and return a BytesIO stream."""
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=(width, height))
    can.setFillColorRGB(*color_rgb)
    can.setFont(font_name, font_size)
    can.drawString(x, y, text)
    can.save()
    packet.seek(0)
    return packet

def add_text_to_pdf(input_path: str, output_path: str, text: str,
                    box: Tuple[float, float, float, float],
                    font_size: int = 24, font_name: str = "Helvetica",
                    color_rgb: Tuple[float, float, float] = (0, 0, 0)):
    """Add text to a PDF file and save it."""
    reader = PdfReader(input_path)
    writer = PdfWriter()

    for page in reader.pages:
        media = page.mediabox
        width = float(media.width)
        height = float(media.height)

        text_width = pdfmetrics.stringWidth(text, font_name, font_size)

        x1, y1, x2, y2 = box
        box_width = x2 - x1
        box_height = y2 - y1

        # Center the text within the bounding box
        x = x1 + (box_width - text_width) / 2
        # Adjust y to be centered vertically. The `y` for drawString is the baseline.
        y = y1 + (box_height - font_size) / 2 + font_size / 4

        overlay_stream = create_overlay_stream(text, width, height, x, y,
                                               font_name=font_name, font_size=font_size,
                                               color_rgb=color_rgb)
        overlay_pdf = PdfReader(overlay_stream)
        overlay_page = overlay_pdf.pages[0]
        page.merge_page(overlay_page)
        writer.add_page(page)

    with open(output_path, "wb") as f_out:
        writer.write(f_out)

def process_pdfs(excel_file, pdf_template, output_folder, name_col, phone_col,
                 box: Tuple[float, float, float, float], font_size=24, font_name="Helvetica"):
    """Process Excel file and create personalized PDFs."""
    try:
        if not os.path.exists(excel_file):
            print(f"File '{excel_file}' does not exist.")
            return
        if not excel_file.lower().endswith('.xlsx'):
            print("Only .xlsx files are supported.")
            return
        if not os.path.exists(pdf_template):
            print(f"PDF template '{pdf_template}' does not exist.")
            return

        # Create output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            print(f"Created output folder: {output_folder}")

        workbook = load_workbook(excel_file)
        sheet = workbook.active

        processed_count = 0
        skipped_count = 0

        for excel_row_idx, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                min_col=min(name_col, phone_col),
                max_col=max(name_col, phone_col),
                values_only=True
            ),
            start=2
        ):
            try:
                name = row[name_col - min(name_col, phone_col)]
                number = row[phone_col - min(name_col, phone_col)]

                if name is None or number is None:
                    print(f"Row {excel_row_idx}: Skipping - missing name or phone number")
                    skipped_count += 1
                    continue

                # Prepare a title-cased display name and clean the phone number for use as filename
                # Only take the first 2 words from the name
                name_parts = str(name).split()
                display_name = ' '.join(name_parts[:2]).title()
                number = str(number).strip().replace(" ", "").replace("+", "").replace("-", "")

                # Create output filename using phone number
                output_filename = f"{number}.pdf"
                output_path = os.path.join(output_folder, output_filename)

                # Add name to PDF (use title-cased name)
                add_text_to_pdf(
                    input_path=pdf_template,
                    output_path=output_path,
                    text=display_name,
                    box=box,
                    font_size=font_size,
                    font_name=font_name,
                    color_rgb=(0, 0, 0)
                )

                print(f"✅ Created: {output_filename} with name: {display_name}")
                processed_count += 1

            except Exception as row_e:
                print(f"❌ Error processing row {excel_row_idx}: {row_e}")
                skipped_count += 1

        print(f"\n{'='*50}")
        print(f"Processing complete!")
        print(f"  ✅ Successfully processed: {processed_count}")
        print(f"  ❌ Skipped/Failed: {skipped_count}")
        print(f"  📁 Output folder: {output_folder}")
        print(f"{'='*50}")

    except Exception as e:
        print(f"Critical error: {e}")

def get_user_input_coordinates() -> Tuple[float, float, float, float]:
    """Prompt user for bounding box coordinates."""
    print("\nEnter the coordinates for the bounding box of the text.")
    print("The origin (0,0) is at the bottom-left corner of the PDF page.")
    while True:
        try:
            x1 = float(input("Enter x1 coordinate: "))
            y1 = float(input("Enter y1 coordinate: "))
            x2 = float(input("Enter x2 coordinate: "))
            y2 = float(input("Enter y2 coordinate: "))
            if x1 >= x2 or y1 >= y2:
                print("Invalid coordinates: x1 must be less than x2 and y1 must be less than y2.")
                continue
            return x1, y1, x2, y2
        except ValueError:
            print("Invalid input. Please enter numbers for the coordinates.")

def get_user_input_font(available_fonts: list) -> str:
    """Prompt user for font name.
    If the font isn't registered, attempt to download it automatically from Google Fonts GitHub.
    """
    print("\nA `FONTS_README.md` file has been created with a list of all available fonts.")
    print("You can copy a font name from there and paste it below.")

    while True:
        choice = input("Enter font name [default: Helvetica]: ").strip()
        if not choice:
            return "Helvetica"

        # Allow entering the font name directly (case-insensitive)
        font_lower = choice.lower()
        for font in available_fonts:
            if font.lower() == font_lower:
                return font

        # Not found locally — attempt to download and register it automatically
        print(f"Font '{choice}' not found locally. Attempting to download and register it...")
        registered_font = download_google_font(choice)
        if registered_font:
            print(f"Using downloaded font: {registered_font}")
            return registered_font

        print(f"Could not download font '{choice}'. Please enter another font name or install the font locally.")

def get_user_input_font_size() -> int:
    """Prompt user for font size."""
    while True:
        size_input = input("Enter font size in points [default: 24]: ").strip()
        if not size_input:
            return 24

        try:
            size = int(size_input)
            if size > 0:
                return size
            else:
                print("Font size must be positive.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def main():
    # Get system fonts at the beginning
    available_fonts = get_system_fonts()
    # Also create the README with the font list
    create_font_readme(available_fonts)

    while True:
        print("\n=== PDF Name Generator ===")
        print("1. Generate personalized PDFs from Excel")
        print("2. Exit")
        choice = input("Enter your choice: ").strip()

        if choice == "1":
            excel_file = input("Enter Excel file name (default: input_data.xlsx): ").strip() or "input_data.xlsx"
            if not excel_file.lower().endswith('.xlsx'):
                excel_file += '.xlsx'

            pdf_template = input("Enter PDF template file name (default: certificate_template.pdf): ").strip() or "certificate_template.pdf"
            if not pdf_template.lower().endswith('.pdf'):
                pdf_template += '.pdf'

            output_folder = input("Enter output folder name (default: output): ").strip() or "output"

            name_col_letter = input("Enter the column letter for the name (default: A): ").strip() or "A"
            phone_col_letter = input("Enter the column letter for the phone number (default: C): ").strip() or "C"

            name_col = col_letter_to_index(name_col_letter)
            phone_col = col_letter_to_index(phone_col_letter)

            if not name_col or not phone_col:
                print("Invalid column letter(s). Please try again.")
                continue

            # Get styling options
            box = get_user_input_coordinates()
            font_name = get_user_input_font(available_fonts)
            font_size = get_user_input_font_size()

            process_pdfs(excel_file, pdf_template, output_folder, name_col, phone_col,
                        box, font_size, font_name)

        elif choice == "2":
            print("Exiting.")
            break
        else:
            print("Invalid choice. Please try again.")

def create_font_readme(font_list):
    """Creates a README.md file with a list of available fonts."""
    content = "# Available Fonts\n\n"
    content += "You can use any of the following fonts in the application. Just copy the font name and paste it when prompted.\n\n"
    content += "| Font Name |\n"
    content += "|-----------|\n"
    for font in font_list:
        content += f"| `{font}` |\n"

    with open("FONTS_README.md", "w", encoding="utf-8") as f:
        f.write(content)


if __name__ == "__main__":
    main()
